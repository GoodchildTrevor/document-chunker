import logging
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from document_chunker.config import NLPConfig
from document_chunker.processing.chunk import preprocess_text

logger = logging.getLogger(__name__)


def excel_to_text(
    file_path: Path,
    log: logging.Logger,
    nlp_config: NLPConfig,
) -> list[dict[str, Any]]:
    """
    Convert each Excel sheet into a list of formatted row strings.
    Returns elements with type='table'.
    """
    elements: list[dict[str, Any]] = []
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        headers = [str(h) if h is not None else "" for h in rows[0]]
        formatted_rows: list[str] = []

        for row in rows[1:]:
            parts = []
            for header, cell in zip(headers, row):
                if cell is not None and str(cell).strip():
                    parts.append(f"{header}: {cell}")
            if parts:
                formatted_rows.append(" | ".join(parts))

        if formatted_rows:
            elements.append({
                "type": "table",
                "content": formatted_rows,
                "_meta": {"sheet": sheet_name},
            })
            log.info("Sheet '%s': %d data rows", sheet_name, len(formatted_rows))

    wb.close()
    return elements


def extract_excel_metadata(file_path: Path, log: logging.Logger) -> dict[str, Any]:
    fallback = datetime(1900, 1, 1, tzinfo=timezone.utc)
    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        props = wb.properties
        creation_date = props.created or fallback
        modification_date = props.modified or fallback
        wb.close()
        return {"creation_date": creation_date, "modification_date": modification_date}
    except Exception as e:
        log.warning("Failed to extract Excel metadata: %s", e)
        return {"creation_date": fallback, "modification_date": fallback}
